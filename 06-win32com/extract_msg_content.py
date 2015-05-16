# -*- encoding: utf8 -*-
# http://www.gossamer-threads.com/lists/python/python/868295

import os, sys
import email
import glob
import re

from win32com.mapi import mapi, mapitags
from win32com.shell import shell, shellcon
from win32com.storagecon import *
import pythoncom

from openpyxl import Workbook
from openpyxl.cell import get_column_letter

# mapitags.PR_BODY, mapitags.PR_SUBJECT, mapitags.PR_TRANSPORT_MESSAGE_HEADERS
def get_data_from_stream (message, prop):
    CHUNK_SIZE = 10000
    stream = message.OpenProperty(prop, pythoncom.IID_IStream, 0, 0)
    text = ""
    while True:
        bytes = stream.read(CHUNK_SIZE)
        if bytes:
            text += bytes
        else:
            break
    return text.decode("utf16")

def parse_msg_file(filepath):
    mapi.MAPIInitialize ((mapi.MAPI_INIT_VERSION, 0))
    storage_flags = STGM_DIRECT | STGM_READ | STGM_SHARE_EXCLUSIVE
    storage = pythoncom.StgOpenStorage (filepath, None, storage_flags, None, 0)
    mapi_session = mapi.OpenIMsgSession ()
    message = mapi.OpenIMsgOnIStg (mapi_session, None, storage, None, 0,
        mapi.MAPI_UNICODE)
    return message

def writeResult(result, result_file):
    wb = Workbook()
    ws = wb.create_sheet()
    ws.title = 'Sheet0'
    header = result[0]

    col_idx = 1
    row_idx = 1
    for v in header:
        ws.cell('%s%d'%(get_column_letter(col_idx), row_idx)).value = v
        col_idx += 1
    
    for row in result[1]:
        col_idx = 1
        row_idx += 1
        for v in row:
            ws.cell('%s%d'%(get_column_letter(col_idx), row_idx)).value = v
            col_idx += 1
    wb.save(filename = result_file)
    
def get_option_values(mail_options):
    e = email.email.message_from_string(mail_options)
    vals = []
    vals.append(e.get('Date'))
    vals.append(e.get('From'))
    vals.append(e.get('To'))
    
    receiveds = e.get_all('Received')
    if len(receiveds) > 0:
        r = receiveds[-1]
        vals.append(r.split('\n')[0].strip())
    else:
        vals.append('')
    vals.append(e.get('Subject'))
    return vals

def extract_headers():
    msg_list = list(glob.glob('*.msg'))
    heads = ['Date', 'From', 'To', 'Received', 'Subject', 'file']
    rows = []
    for n in msg_list:
        r = get_option_values(get_data_from_stream(parse_msg_file(n), mapitags.PR_TRANSPORT_MESSAGE_HEADERS))
        r.append(n)
        rows.append(r)
    writeResult((heads, rows), 'headers.xlsx')

def extract_addr_from_content():
    email_rex = re.compile('[a-zA-Z0-9_\'\\-]+(\\.[a-zA-Z0-9_\'\\-]+)*@[a-zA-Z0-9_\'\\-]+(\\.[a-zA-Z0-9_\'\\-]+)*[\\.][a-zA-Z]{2,3}([\\.][a-zA-Z]{2})?', re.U)
    
    o = open('out.txt', 'wb')
    msg_list = list(glob.glob('*.msg'))
    for n in msg_list:
        content = get_data_from_stream(parse_msg_file(n), mapitags.PR_BODY)
        for m in email_rex.finditer(content):
            o.write(m.group(0))
            o.write('\n')
        #addr_list = email_rex.findall(content) # not work
    o.close()

if __name__ == '__main__':
    #extract_headers()
    extract_addr_from_content()
